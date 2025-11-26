import os
import sys
import json
import time
import random
import uuid
import hashlib
import threading
import sqlite3
import asyncio
from datetime import datetime, timedelta
from typing import Dict, Optional, Tuple, List
import requests
from requests import Session
from requests.utils import dict_from_cookiejar
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from faker import Faker
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, CallbackQueryHandler, MessageHandler, filters, ContextTypes, ConversationHandler
import logging
from urllib.parse import quote, urlencode
import base64
import re

# ==================== CONFIG ====================
BOT_TOKEN = "8384781635:AAG4BPHB_zQA5ld7U-vzr6c72GIgApysTE0"
OWNER_ID = 8244615483

# Paths
STORAGE_DIR = "/storage/emulated/0/bot_data"
os.makedirs(STORAGE_DIR, exist_ok=True)
DB_PATH = os.path.join(STORAGE_DIR, "fb_creator_v3.db")
XLSX_FILE = os.path.join(STORAGE_DIR, "Acc_Created_V3.xlsx")
TXT_FILE = os.path.join(STORAGE_DIR, "Acc_created_V3.txt")
PROSES_PATH = "/storage/emulated/0/proses.png"

# ==================== LOGGING ====================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ==================== FAKERS ====================
fake_en = Faker('en_US')
fake_id = Faker('id_ID')
fake_es = Faker('es_ES')

# ==================== STATES ====================
GENDER_SELECT, METHOD_SELECT, EMAIL_INPUT, PHONE_INPUT, PASSWORD_INPUT, OWNER_ADD_DOMAIN, DOMAIN_SELECT, LOCALE_SELECT, COUNTRY_SELECT = range(9)

# ==================== GLOBALS ====================
creation_semaphore = threading.Semaphore(50)
user_sessions: Dict[int, Session] = {}
active_tasks: Dict[int, bool] = {}

# ==================== TEMP DOMAINS (UPDATED) ====================
TEMP_DOMAINS = [
    "08058383859.com",
    "09063999201.com",
    "keishichootsukakeisatsusho.com",
    "unko.delivery",
    "nyaacarding.group",
    "paicharm.tokyo",
    "hamutan86.jp"
]

from faker import Faker

def safe_faker(locale: str):
    try:
        return Faker(locale)
    except:
        return Faker('en_US')

COUNTRY_FAKERS = {
    'US': safe_faker('en_US'),
    'UK': safe_faker('en_GB'),
    'CA': safe_faker('en_CA'),
    'AU': safe_faker('en_AU'),
    'ID': safe_faker('id_ID'),
    'IN': safe_faker('en_US'),
    'PH': safe_faker('en_US'),
    'MY': safe_faker('en_US'),
    'SG': safe_faker('en_US'),
    'TH': safe_faker('th_TH'),
    'VN': safe_faker('vi_VN'),
    'JP': safe_faker('ja_JP'),
    'KR': safe_faker('ko_KR'),
    'CN': safe_faker('zh_CN'),
    'TW': safe_faker('zh_TW'),
    'HK': safe_faker('en_US'),
    'BR': safe_faker('pt_BR'),
    'MX': safe_faker('es_MX'),
    'AR': safe_faker('en_US'),
    'CL': safe_faker('en_US'),
    'CO': safe_faker('en_US'),
    'ES': safe_faker('es_ES'),
    'FR': safe_faker('fr_FR'),
    'DE': safe_faker('de_DE'),
    'IT': safe_faker('it_IT'),
    'NL': safe_faker('nl_NL'),
    'PL': safe_faker('pl_PL'),
    'RU': safe_faker('ru_RU'),
    'TR': safe_faker('tr_TR'),
    'SA': safe_faker('ar_SA'),
    'AE': safe_faker('en_US'),
    'EG': safe_faker('ar_EG'),
    'ZA': safe_faker('en_US'),
    'NG': safe_faker('en_US'),
    'KE': safe_faker('en_US'),
    'GH': safe_faker('en_US'),
    'ET': safe_faker('en_US'),
    'TZ': safe_faker('en_US'),
    'UG': safe_faker('en_US'),
    'MA': safe_faker('fr_FR'),
    'DZ': safe_faker('ar_EG'),
    'TN': safe_faker('ar_TN'),
    'LY': safe_faker('ar_EG'),
    'SD': safe_faker('ar_EG'),
    'IQ': safe_faker('ar_IQ'),
    'SY': safe_faker('ar_SY'),
    'JO': safe_faker('ar_JO'),
    'LB': safe_faker('ar_LB'),
    'KW': safe_faker('ar_KW'),
    'QA': safe_faker('ar_QA'),
    'BH': safe_faker('ar_EG'),
    'OM': safe_faker('ar_OM'),
    'YE': safe_faker('ar_YE'),
}

# ==================== COUNTRY PHONE CONFIGS ====================
COUNTRY_PHONE_CONFIG = {
    'US': {'code': '1', 'length': 10, 'prefixes': ['201', '202', '203', '212', '213', '214', '215', '216', '217', '218']},
    'UK': {'code': '44', 'length': 10, 'prefixes': ['7400', '7401', '7402', '7403', '7404', '7405', '7500', '7501', '7502', '7503']},
    'CA': {'code': '1', 'length': 10, 'prefixes': ['204', '226', '236', '249', '250', '289', '306', '343', '365', '367']},
    'AU': {'code': '61', 'length': 9, 'prefixes': ['400', '401', '402', '403', '404', '405', '406', '407', '408', '409']},
    'ID': {'code': '62', 'length': 10, 'prefixes': ['811', '812', '813', '814', '815', '816', '817', '818', '819', '821']},
    'IN': {'code': '91', 'length': 10, 'prefixes': ['70', '71', '72', '73', '74', '75', '76', '77', '78', '79', '80', '81', '82', '83', '84', '85', '86', '87', '88', '89', '90', '91', '92', '93', '94', '95', '96', '97', '98', '99']},
    'PH': {'code': '63', 'length': 10, 'prefixes': ['905', '906', '907', '908', '909', '910', '911', '912', '913', '914']},
    'MY': {'code': '60', 'length': 9, 'prefixes': ['10', '11', '12', '13', '14', '15', '16', '17', '18', '19']},
    'SG': {'code': '65', 'length': 8, 'prefixes': ['8000', '8001', '8002', '8003', '8004', '8100', '8101', '8102', '8103', '8104']},
    'TH': {'code': '66', 'length': 9, 'prefixes': ['80', '81', '82', '83', '84', '85', '86', '87', '88', '89']},
    'VN': {'code': '84', 'length': 9, 'prefixes': ['70', '76', '77', '78', '79', '81', '82', '83', '84', '85', '86', '88', '89', '90', '91', '92', '93', '94', '96', '97', '98', '99']},
    'JP': {'code': '81', 'length': 10, 'prefixes': ['70', '80', '90']},
    'KR': {'code': '82', 'length': 9, 'prefixes': ['10']},
    'CN': {'code': '86', 'length': 11, 'prefixes': ['130', '131', '132', '133', '134', '135', '136', '137', '138', '139', '147', '150', '151', '152', '153', '155', '156', '157', '158', '159', '178', '180', '181', '182', '183', '184', '185', '186', '187', '188', '189']},
    'BR': {'code': '55', 'length': 11, 'prefixes': ['11', '12', '13', '14', '15', '16', '17', '18', '19', '21']},
    'MX': {'code': '52', 'length': 10, 'prefixes': ['55', '33', '81']},
    'ES': {'code': '34', 'length': 9, 'prefixes': ['6', '7']},
    'FR': {'code': '33', 'length': 9, 'prefixes': ['6', '7']},
    'DE': {'code': '49', 'length': 10, 'prefixes': ['151', '152', '157', '159', '160', '162', '163', '170', '171', '172', '173', '174', '175', '176', '177', '178', '179']},
    'IT': {'code': '39', 'length': 10, 'prefixes': ['320', '323', '327', '328', '329', '330', '331', '333', '334', '335', '336', '337', '338', '339', '340', '342', '343', '344', '345', '346', '347', '348', '349']},
    'NL': {'code': '31', 'length': 9, 'prefixes': ['6']},
    'RU': {'code': '7', 'length': 10, 'prefixes': ['900', '901', '902', '903', '904', '905', '906', '908', '909', '910', '911', '912', '913', '914', '915', '916', '917', '918', '919', '920', '921', '922', '923', '924', '925', '926', '927', '928', '929']},
    'TR': {'code': '90', 'length': 10, 'prefixes': ['501', '505', '506', '507', '530', '531', '532', '533', '534', '535', '536', '537', '538', '539', '540', '541', '542', '543', '544', '545', '546', '547', '548', '549']},
    'SA': {'code': '966', 'length': 9, 'prefixes': ['50', '51', '52', '53', '54', '55', '56', '57', '58', '59']},
    'AE': {'code': '971', 'length': 9, 'prefixes': ['50', '52', '54', '55', '56', '58']},
    'EG': {'code': '20', 'length': 10, 'prefixes': ['10', '11', '12', '15']},
    'ZA': {'code': '27', 'length': 9, 'prefixes': ['60', '61', '62', '63', '64', '65', '66', '67', '68', '69', '70', '71', '72', '73', '74', '76', '77', '78', '79', '81', '82', '83', '84']},
    'NG': {'code': '234', 'length': 10, 'prefixes': ['701', '702', '703', '704', '705', '706', '707', '708', '709', '802', '803', '804', '805', '806', '807', '808', '809', '810', '811', '812', '813', '814', '815', '816', '817', '818']},
    'KE': {'code': '254', 'length': 9, 'prefixes': ['70', '71', '72', '74', '75', '76', '78', '79']},
    'AR': {'code': '54', 'length': 10, 'prefixes': ['11', '15']},
    'CL': {'code': '56', 'length': 9, 'prefixes': ['9']},
    'CO': {'code': '57', 'length': 10, 'prefixes': ['300', '301', '302', '303', '304', '305', '310', '311', '312', '313', '314', '315', '316', '317', '318', '319', '320', '321', '322', '323', '324', '350', '351']},
    'PL': {'code': '48', 'length': 9, 'prefixes': ['45', '50', '51', '53', '57', '60', '66', '69', '72', '73', '78', '79', '88']},
}

# ==================== FACEBOOK LOCALES ====================
FB_LOCALES = {
    'en': {'code': 'en_US', 'name': '🇺🇸 English (US)', 'lang': 'English'},
    'id': {'code': 'id_ID', 'name': '🇮🇩 Bahasa Indonesia', 'lang': 'Indonesian'},
    'es': {'code': 'es_LA', 'name': '🇪🇸 Español (Latinoamérica)', 'lang': 'Spanish'}
}

# ==================== DEVICE INFO GENERATORS ====================
def get_random_device():
    devices = [
        {
            'model': 'Samsung-SM-S918B',
            'android': '14',
            'build': 'UP1A.231005.007',
            'brand': 'samsung',
            'device': 's918b',
            'fingerprint': 'samsung/s918bxx/s918b:14/UP1A.231005.007/S918BXXU2BWK4:user/release-keys'
        },
        {
            'model': 'Xiaomi-2210132G',
            'android': '13',
            'build': 'TKQ1.221114.001',
            'brand': 'xiaomi',
            'device': 'marble',
            'fingerprint': 'Xiaomi/marble_global/marble:13/TKQ1.221114.001/V14.0.8.0.TMRMIXM:user/release-keys'
        },
        {
            'model': 'OnePlus-CPH2451',
            'android': '14',
            'build': 'UKQ1.230924.001',
            'brand': 'oneplus',
            'device': 'OP594DL1',
            'fingerprint': 'OnePlus/CPH2451IND/OP594DL1:14/UKQ1.230924.001/U.R4T3.16e6e87-4ae-6b6:user/release-keys'
        },
        {
            'model': 'OPPO-CPH2207',
            'android': '13',
            'build': 'TP1A.220905.001',
            'brand': 'oppo',
            'device': 'OP4BA2L1',
            'fingerprint': 'OPPO/CPH2207/OP4BA2L1:13/TP1A.220905.001/R.14e1e8c-6b8:user/release-keys'
        },
        {
            'model': 'vivo-V2203',
            'android': '13',
            'build': 'TP1A.220624.014',
            'brand': 'vivo',
            'device': 'PD2203',
            'fingerprint': 'vivo/V2203/PD2203:13/TP1A.220624.014/compiler08241734:user/release-keys'
        },
        {
            'model': 'Google-Pixel-7-Pro',
            'android': '14',
            'build': 'UP1A.231105.001',
            'brand': 'google',
            'device': 'cheetah',
            'fingerprint': 'google/cheetah/cheetah:14/UP1A.231105.001/10817346:user/release-keys'
        },
        {
            'model': 'Samsung-SM-A546E',
            'android': '14',
            'build': 'UP1A.231005.007',
            'brand': 'samsung',
            'device': 'a54x',
            'fingerprint': 'samsung/a54xdx/a54x:14/UP1A.231005.007/A546EDXU3CXC1:user/release-keys'
        }
    ]
    
    device = random.choice(devices)
    device['width'] = random.choice([1080, 1440, 2400])
    device['height'] = random.choice([2340, 3088, 3200])
    device['dpi'] = random.choice([420, 480, 560, 640])
    
    return device

def generate_user_agent(device, fb_app_version="484.0.0.14.106"):
    chrome_version = f"{random.randint(120, 142)}.0.{random.randint(7000, 7999)}.{random.randint(100, 200)}"
    webkit_version = "537.36"
    
    ua = f"Mozilla/5.0 (Linux; Android {device['android']}; {device['model']} Build/{device['build']}; wv) "
    ua += f"AppleWebKit/{webkit_version} (KHTML, like Gecko) Version/4.0 "
    ua += f"Chrome/{chrome_version} Mobile Safari/{webkit_version}"
    ua += f"[FBAN/EMA;FBLC/en_US;FBAV/{fb_app_version};]"
    
    return ua

def generate_advanced_headers(device, locale='en_US'):
    ua = generate_user_agent(device)
    
    headers = {
        'Host': 'm.facebook.com',
        'User-Agent': ua,
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
        'Accept-Language': f'{locale.replace("_", "-")},en-US;q=0.9,en;q=0.8',
        'Accept-Encoding': 'gzip, deflate, br',
        'Connection': 'keep-alive',
        'Upgrade-Insecure-Requests': '1',
        'Sec-Fetch-Dest': 'document',
        'Sec-Fetch-Mode': 'navigate',
        'Sec-Fetch-Site': 'none',
        'Sec-Fetch-User': '?1',
        'Cache-Control': 'max-age=0',
        'DNT': '1',
        'X-Requested-With': 'com.facebook.lite',
        'X-FB-Connection-Type': random.choice(['WIFI', 'mobile.LTE', '4G']),
        'X-FB-Connection-Quality': 'EXCELLENT',
        'X-FB-Device': device['model'],
        'X-FB-Device-Group': str(random.randint(5000, 9999)),
        'X-FB-Net-HNI': str(random.randint(40000, 52000)),
        'X-FB-SIM-HNI': str(random.randint(40000, 52000)),
        'X-FB-HTTP-Engine': 'Liger',
        'X-FB-Client-IP': 'True',
        'X-FB-Server-Cluster': 'True',
    }
    
    return headers

def generate_random_phone_by_country(country_code='ID'):
    if country_code not in COUNTRY_PHONE_CONFIG:
        country_code = 'ID'
    
    config = COUNTRY_PHONE_CONFIG[country_code]
    prefix = random.choice(config['prefixes'])
    remaining = config['length'] - len(prefix)
    
    number = prefix + ''.join([str(random.randint(0, 9)) for _ in range(remaining)])
    full_number = config['code'] + number
    
    return full_number

def generate_random_hotmail():
    prefixes = ['', 'the', 'real', 'official', 'my', 'its']
    suffixes = ['', str(random.randint(1, 999)), str(random.randint(1990, 2005))]
    
    username = random.choice(prefixes) + fake_en.user_name().replace('.', '').replace('_', '') + random.choice(suffixes)
    domains = ['@hotmail.com', '@outlook.com', '@live.com']
    
    return username.lower() + random.choice(domains)

def generate_random_gmail():
    username = fake_en.user_name().replace('.', '').replace('_', '') + str(random.randint(100, 9999))
    return f"{username.lower()}@gmail.com"

# ==================== DATABASE ====================
def init_db():
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()

        c.execute('''CREATE TABLE IF NOT EXISTS users (
            user_id INTEGER PRIMARY KEY,
            username TEXT,
            language TEXT DEFAULT 'en',
            gender TEXT DEFAULT 'male',
            approved INTEGER DEFAULT 0,
            banned INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')

        ensure_user_columns(conn)

        c.execute('''CREATE TABLE IF NOT EXISTS domains (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            domain TEXT UNIQUE
        )''')

        for domain in TEMP_DOMAINS:
            c.execute("INSERT OR IGNORE INTO domains (domain) VALUES (?)", (domain,))

        c.execute("""
            INSERT OR REPLACE INTO users
            (user_id, username, language, gender, approved, banned)
            VALUES (?, ?, ?, ?, 1, 0)
        """, (OWNER_ID, "Owner", "en", "male"))

        conn.commit()
        conn.close()
        logger.info("🔥 Database initialized - VoonzyV3")
    except Exception as e:
        logger.error(f"DB init error: {e}")

def ensure_user_columns(conn):
    try:
        c = conn.cursor()
        c.execute("PRAGMA table_info(users);")
        existing = [row[1] for row in c.fetchall()]

        if "gender" not in existing:
            c.execute("ALTER TABLE users ADD COLUMN gender TEXT DEFAULT 'male'")
        if "approved" not in existing:
            c.execute("ALTER TABLE users ADD COLUMN approved INTEGER DEFAULT 0")
        if "banned" not in existing:
            c.execute("ALTER TABLE users ADD COLUMN banned INTEGER DEFAULT 0")

        conn.commit()
    except Exception as e:
        logger.debug(f"ensure_user_columns warning: {e}")

def get_user(user_id: int) -> Optional[Dict]:
    try:
        conn = sqlite3.connect(DB_PATH)
        ensure_user_columns(conn)

        c = conn.cursor()
        c.execute("SELECT user_id, username, language, gender, approved, banned FROM users WHERE user_id=?", (user_id,))
        row = c.fetchone()
        conn.close()

        if row:
            return {
                "user_id": row[0],
                "username": row[1] or "User",
                "language": row[2] or "en",
                "gender": row[3] or "male",
                "approved": bool(row[4]),
                "banned": bool(row[5])
            }
    except Exception as e:
        logger.error(f"Get user error: {e}")
    return None

def add_or_update_user(user_id: int, username: str, language: str = "en", gender: str = "male"):
    try:
        conn = sqlite3.connect(DB_PATH)
        ensure_user_columns(conn)
        c = conn.cursor()

        c.execute("""
            INSERT OR IGNORE INTO users (user_id, username, language, gender, approved, banned)
            VALUES (?, ?, ?, ?, 0, 0)
        """, (user_id, username, language, gender))

        c.execute("""
            UPDATE users SET username = ?, language = ?, gender = ? WHERE user_id = ?
        """, (username, language, gender, user_id))

        conn.commit()
        conn.close()
        return True
    except Exception as e:
        logger.error(f"Add user error: {e}")
        return False

def set_approved(user_id: int, approved: bool):
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("UPDATE users SET approved=? WHERE user_id=?", (approved, user_id))
        conn.commit()
        conn.close()
    except Exception as e:
        logger.error(f"Set approved error: {e}")

def set_banned(user_id: int, banned: bool):
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("UPDATE users SET banned=? WHERE user_id=?", (banned, user_id))
        conn.commit()
        conn.close()
    except Exception as e:
        logger.error(f"Set banned error: {e}")

def get_pending_users() -> List[Dict]:
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("SELECT user_id, username FROM users WHERE approved=0 AND banned=0")
        rows = c.fetchall()
        conn.close()
        return [{"user_id": r[0], "username": r[1] or "User"} for r in rows]
    except:
        return []

def get_all_users() -> List[Dict]:
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("SELECT user_id, username, approved, banned FROM users WHERE user_id != ?", (OWNER_ID,))
        rows = c.fetchall()
        conn.close()
        return [{"user_id": r[0], "username": r[1] or "User", "approved": bool(r[2]), "banned": bool(r[3])} for r in rows]
    except:
        return []

def get_domains() -> List[str]:
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("SELECT domain FROM domains")
        domains = [row[0] for row in c.fetchall()]
        conn.close()
        return domains if domains else TEMP_DOMAINS
    except:
        return TEMP_DOMAINS

def add_domain(domain: str) -> bool:
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("INSERT INTO domains (domain) VALUES (?)", (domain,))
        conn.commit()
        conn.close()
        return True
    except:
        return False

def delete_domain(domain: str):
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("DELETE FROM domains WHERE domain=?", (domain,))
        conn.commit()
        conn.close()
    except:
        pass

# ==================== UTILITY FUNCTIONS ====================
def get_fake_name(lang: str, gender: str, country_code: str = 'US') -> Tuple[str, str]:
    if country_code in COUNTRY_FAKERS:
        fake = COUNTRY_FAKERS[country_code]
    elif lang == 'es':
        fake = fake_es
    elif lang == 'id':
        fake = fake_id
    else:
        fake = fake_en
    
    first = fake.first_name_male() if gender == 'male' else fake.first_name_female()
    return first, fake.last_name()

def generate_password(base: str = "Password"):
    if len(base) >= 6:
        return base
    return base + str(random.randint(100000, 999999))

def generate_temp_email(domain: Optional[str] = None):
    user_part = fake_en.user_name() + str(random.randint(100, 999))
    if domain is None:
        domain = random.choice(get_domains())
    return f"{user_part}@{domain}", user_part, domain

def get_user_session(user_id: int) -> Session:
    if user_id not in user_sessions:
        user_sessions[user_id] = Session()
    return user_sessions[user_id]

def clear_session(user_id: int):
    if user_id in user_sessions:
        try:
            user_sessions[user_id].close()
        except:
            pass
        del user_sessions[user_id]
    logger.info(f"💀 Session cleared for user {user_id}")

def save_to_xlsx(filename: str, data: List):
    try:
        header_columns = ['NAME', 'USERNAME', 'PASSWORD', 'PROFILE', 'COOKIES']
        if os.path.exists(filename):
            try:
                wb = load_workbook(filename)
                ws = wb.active
            except:
                wb = Workbook()
                ws = wb.active
                ws.append(header_columns)
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(header_columns)
        ws.append(data)
        wb.save(filename)
    except Exception as e:
        logger.error(f"XLSX save error: {e}")

def save_to_txt(filename: str, data: List):
    try:
        with open(filename, "a", encoding="utf-8") as f:
            f.write("|".join(data) + "\n")
    except Exception as e:
        logger.error(f"TXT save error: {e}")

def get_full_cookies(session: Session) -> str:
    try:
        cookies_dict = dict_from_cookiejar(session.cookies)
        cookies_str = "; ".join([f"{k}={v}" for k, v in cookies_dict.items()])
        return cookies_str if cookies_str else "NO_COOKIES"
    except:
        return "NO_COOKIES"

# ==================== ADVANCED ANTI-CHECKPOINT TECHNIQUES ====================
def generate_advanced_device_signature(device):
    timestamp = int(time.time())
    random_id = str(uuid.uuid4())
    
    signature_data = f"{device['model']}{device['android']}{timestamp}{random_id}"
    signature = hashlib.md5(signature_data.encode()).hexdigest()
    
    return {
        'device_id': str(uuid.uuid4()),
        'family_device_id': str(uuid.uuid4()),
        'device_group': str(random.randint(5000, 9999)),
        'advertising_id': str(uuid.uuid4()),
        'session_id': str(uuid.uuid4()),
        'signature': signature
    }

def simulate_realistic_delays():
    base_delay = random.uniform(2.0, 4.5)
    variance = random.uniform(-0.3, 1.0)
    time.sleep(max(1.0, base_delay + variance))

def generate_fb_dtsg(session: Session, url: str) -> str:
    try:
        response = session.get(url, timeout=30)
        match = re.search(r'"DTSGInitialData",\[\],{"token":"([^"]+)"', response.text)
        if match:
            return match.group(1)
        
        match = re.search(r'name="fb_dtsg" value="([^"]+)"', response.text)
        if match:
            return match.group(1)
        
        match = re.search(r'"dtsg":{"token":"([^"]+)"', response.text)
        if match:
            return match.group(1)
    except:
        pass
    return ""

def generate_jazoest(fb_dtsg: str) -> str:
    try:
        result = 0
        for char in fb_dtsg:
            result += ord(char)
        return f"2{result}"
    except:
        return "22581"

# ==================== HARDCORE FB ACCOUNT CREATION ====================
def create_facebook_account_hardcore(
    user_id: int,
    method: int,
    email: Optional[str],
    phone: Optional[str],
    password: str,
    lang: str,
    gender: str,
    selected_domain: Optional[str] = None,
    fb_locale: str = 'en_US',
    country_code: str = 'US',
    message_id: int = None,
    context = None
) -> Dict:

    def schedule_coroutine(coro):
        try:
            loop = None
            try:
                loop = asyncio.get_running_loop()
            except RuntimeError:
                try:
                    loop = asyncio.get_event_loop()
                except Exception:
                    loop = None

            if loop and loop.is_running():
                asyncio.run_coroutine_threadsafe(coro, loop)
                return True
        except Exception as ex:
            logger.debug(f"Schedule coroutine failed: {ex}")
        return False

    def safe_edit_message(chat_id, message_id, text, reply_markup=None):
        if not context:
            return False
        try:
            coro = context.bot.edit_message_text(chat_id=chat_id, message_id=message_id, text=text, reply_markup=reply_markup)
            return schedule_coroutine(coro)
        except Exception as e:
            logger.debug(f"safe_edit_message prepare failed: {e}")
            return False

    with creation_semaphore:
        max_retry_attempts = 5

        for retry_count in range(max_retry_attempts):
            try:
                active_tasks[user_id] = True

                clear_session(user_id)
                session = get_user_session(user_id)

                device = get_random_device()
                device_sig = generate_advanced_device_signature(device)

                first, last = get_fake_name(lang, gender, country_code)
                full_name = f"{first} {last}"

                birth_day = random.randint(1, 28)
                birth_month = random.randint(1, 12)
                birth_year = random.randint(1988, 2005)

                final_username = ""
                temp_link = ""
                needs_email_change = False
                needs_phone_change = False
                temp_email_for_change = None
                phone_for_change = None

                if method == 1:
                    final_username = phone if phone else generate_random_phone_by_country(country_code)
                    needs_email_change = True
                elif method == 2:
                    final_username = email
                elif method == 3:
                    final_username = phone if phone else generate_random_phone_by_country(country_code)
                elif method == 4:
                    final_username = phone if phone else generate_random_phone_by_country(country_code)
                    needs_email_change = True
                    temp_email, user_part, domain = generate_temp_email(selected_domain)
                    temp_email_for_change = temp_email
                    temp_link = f"https://mail.paicha.cloud/mailbox/{user_part}%40{domain}"
                elif method == 5:
                    temp_email, user_part, domain = generate_temp_email(selected_domain)
                    final_username = temp_email
                    temp_link = f"https://mail.paicha.cloud/mailbox/{user_part}%40{domain}"
                elif method == 6:
                    gmail = generate_random_gmail()
                    final_username = gmail
                    needs_email_change = True
                    temp_email, user_part, domain = generate_temp_email(selected_domain)
                    temp_email_for_change = temp_email
                    temp_link = f"https://mail.paicha.cloud/mailbox/{user_part}%40{domain}"
                elif method == 7:
                    final_username = generate_random_phone_by_country(country_code)
                    needs_email_change = True
                    temp_email_for_change = email if email else generate_random_gmail()
                elif method == 8:
                    final_username = generate_random_hotmail()
                    needs_phone_change = True
                    phone_for_change = generate_random_phone_by_country(country_code)
                elif method == 9:
                    final_username = generate_random_hotmail()
                    needs_email_change = True
                    temp_email, user_part, domain = generate_temp_email(selected_domain)
                    temp_email_for_change = temp_email
                    temp_link = f"https://mail.paicha.cloud/mailbox/{user_part}%40{domain}"
                elif method == 10:
                    final_username = generate_random_phone_by_country(country_code)
                    needs_email_change = True
                    temp_email, user_part, domain = generate_temp_email(selected_domain)
                    temp_email_for_change = temp_email
                    temp_link = f"https://mail.paicha.cloud/mailbox/{user_part}%40{domain}"
                elif method == 11:
                    final_username = generate_random_phone_by_country(country_code)
                    needs_phone_change = True
                    phone_for_change = generate_random_phone_by_country(country_code)

                if message_id and context:
                    try:
                        text = (
                            f"✨ 𝐕𝐎𝐎𝐍𝐙𝐘 𝐕𝟑 𝐇𝐀𝐑𝐃𝐂𝐎𝐑𝐄\n\n"
                            f"🔄 Percobaan {retry_count + 1} dari {max_retry_attempts}\n"
                            f"👤 {full_name}\n"
                            f"📧 {final_username}\n"
                            f"🌍 {fb_locale}\n"
                            f"📱 {device['model']}\n\n"
                            f"⚡ Menembus sistem keamanan"
                        )
                        safe_edit_message(user_id, message_id, text)
                    except Exception:
                        pass

                logger.info(f"💀 VoonzyV3 Attempt {retry_count + 1}/{max_retry_attempts}: {full_name} | {final_username} | Locale: {fb_locale}")

                base_url = f"https://m.facebook.com/reg/?locale={fb_locale}"

                headers = generate_advanced_headers(device, fb_locale)

                form = None
                form_attempts = 0
                fb_dtsg = ""
                jazoest = ""

                while not form and form_attempts < 5:
                    form_attempts += 1
                    try:
                        logger.info(f"⚡ Fetching form (attempt {form_attempts})...")
                        simulate_realistic_delays()
                        response = session.get(base_url, headers=headers, timeout=60)
                        soup = BeautifulSoup(response.text, "html.parser")
                        form = soup.find("form")

                        if form:
                            logger.info("✅ Form fetched successfully")
                            fb_dtsg = generate_fb_dtsg(session, base_url)
                            if fb_dtsg:
                                jazoest = generate_jazoest(fb_dtsg)
                                logger.info(f"🔐 Security tokens extracted")
                            time.sleep(random.uniform(1.5, 2.5))
                            break
                        time.sleep(3)
                    except Exception as e:
                        logger.warning(f"⚠️ Form fetch error: {e}")
                        time.sleep(3)

                if not form:
                    if retry_count < max_retry_attempts - 1:
                        logger.warning(f"⚠️ Form fetch failed, retrying...")
                        continue
                    else:
                        return {
                            "success": False,
                            "error": "Could not load registration form after all retries"
                        }

                reg_data = {
                    "firstname": first,
                    "lastname": last,
                    "birthday_day": str(birth_day),
                    "birthday_month": str(birth_month),
                    "birthday_year": str(birth_year),
                    "reg_email__": final_username,
                    "sex": "1" if gender == 'male' else "2",
                    "encpass": password,
                    "submit": "Sign Up",
                    "locale": fb_locale,
                    "client_country_code": country_code,
                }

                if fb_dtsg:
                    reg_data['fb_dtsg'] = fb_dtsg
                if jazoest:
                    reg_data['jazoest'] = jazoest

                reg_data['device_id'] = device_sig['device_id']
                reg_data['family_device_id'] = device_sig['family_device_id']

                if form:
                    action_url = requests.compat.urljoin(base_url, form.get("action", base_url))
                    for inp in form.find_all("input"):
                        if inp.has_attr("name") and inp["name"] not in reg_data:
                            reg_data[inp["name"]] = inp.get("value", "")

                    post_headers = headers.copy()
                    post_headers['Content-Type'] = 'application/x-www-form-urlencoded'
                    post_headers['Origin'] = 'https://m.facebook.com'
                    post_headers['Referer'] = base_url

                    try:
                        logger.info("💀 Submitting registration with hardcore bypass...")
                        simulate_realistic_delays()
                        reg_response = session.post(action_url, headers=post_headers, data=reg_data, timeout=60, allow_redirects=True)
                        logger.info(f"📡 Response status: {reg_response.status_code}")
                    except requests.exceptions.RequestException as e:
                        logger.warning(f"⚠️ Post exception (might be normal): {e}")
                        pass

                registration_success = "c_user" in session.cookies

                if not registration_success:
                    logger.warning(f"⚠️ No c_user cookie on attempt {retry_count + 1}")

                    if retry_count < max_retry_attempts - 1:
                        if method in [1, 3, 4, 7, 10, 11]:
                            phone = generate_random_phone_by_country(country_code)
                            logger.info(f"🎲 Generated new phone: {phone}")
                        elif method in [8, 9]:
                            email = generate_random_hotmail()
                            logger.info(f"🎲 Generated new hotmail: {email}")

                        if message_id and context:
                            try:
                                text = (
                                    f"✨ 𝐕𝐎𝐎𝐍𝐙𝐘 𝐕𝟑 𝐇𝐀𝐑𝐃𝐂𝐎𝐑𝐄\n\n"
                                    f"⚠️ Mencoba ulang {retry_count + 1} dari {max_retry_attempts}\n"
                                    f"🔄 Cookie tidak terdeteksi\n"
                                    f"♻️ Membuat sesi baru\n"
                                    f"🎲 Generate kredensial baru\n\n"
                                    f"⚡ Tidak pernah menyerah"
                                )
                                safe_edit_message(user_id, message_id, text)
                            except:
                                pass

                        time.sleep(random.uniform(3, 6))
                        continue
                    else:
                        return {
                            "success": False,
                            "error": "Registration failed after all attempts. No c_user cookie detected."
                        }

                logger.info("✅ REGISTRATION SUCCESSFUL - c_user found!")

                uid = session.cookies.get("c_user", "UNKNOWN")
                profile_link = f"https://www.facebook.com/profile.php?id={uid}"

                if needs_email_change and registration_success:
                    new_email = email if method == 1 else temp_email_for_change
                    logger.info(f"📧 Changing to email: {new_email}")

                    try:
                        simulate_realistic_delays()

                        change_email_url = f"https://m.facebook.com/changeemail/?locale={fb_locale}"
                        response = session.get(change_email_url, headers=headers, timeout=60)
                        soup = BeautifulSoup(response.text, "html.parser")
                        form = soup.find("form")

                        if form:
                            action_url = requests.compat.urljoin(change_email_url, form.get("action", change_email_url))
                            data = {}
                            for inp in form.find_all("input"):
                                if inp.has_attr("name"):
                                    data[inp["name"]] = inp.get("value", "")

                            data["new"] = new_email
                            data["submit"] = "Add"

                            if fb_dtsg:
                                data['fb_dtsg'] = fb_dtsg

                            session.post(action_url, headers=headers, data=data, timeout=60)
                            time.sleep(2)
                            check_response = session.get(change_email_url, headers=headers, timeout=60)
                            if "email" in check_response.text.lower():
                                logger.info("✅ Email changed successfully")
                                final_username = new_email
                    except Exception as e:
                        logger.error(f"💀 Email change error: {e}")

                if needs_phone_change and registration_success:
                    new_phone = phone_for_change
                    logger.info(f"📱 Adding phone: {new_phone}")

                    try:
                        simulate_realistic_delays()
                    except Exception as e:
                        logger.error(f"💀 Phone change error: {e}")

                full_cookies = get_full_cookies(session)

                data_to_save = [full_name, final_username, password, profile_link, full_cookies]
                save_to_xlsx(XLSX_FILE, data_to_save)
                save_to_txt(TXT_FILE, data_to_save)

                logger.info(f"💀 VoonzyV3 - Account created: {uid}")

                clear_session(user_id)
                active_tasks.pop(user_id, None)

                return {
                    "success": True,
                    "name": full_name,
                    "username": final_username,
                    "password": password,
                    "uid": uid,
                    "profile": profile_link,
                    "cookies": full_cookies,
                    "temp_link": temp_link,
                    "locale": fb_locale,
                    "country": country_code
                }

            except Exception as e:
                logger.error(f"💀 Account creation error on attempt {retry_count + 1}: {e}")
                if retry_count < max_retry_attempts - 1:
                    time.sleep(random.uniform(4, 8))
                    continue
                else:
                    clear_session(user_id)
                    active_tasks.pop(user_id, None)
                    return {
                        "success": False,
                        "error": f"Failed after {max_retry_attempts} attempts: {str(e)}"
                    }

        clear_session(user_id)
        active_tasks.pop(user_id, None)
        return {
            "success": False,
            "error": "Maximum retry attempts reached"
        }

# ==================== LANGUAGES (V3 HARDCORE - UPDATED UI) ====================
LANGUAGES = {
    'en': {
        'welcome': '✨ 𝐅𝐁 𝐂𝐑𝐄𝐀𝐓𝐎𝐑 𝐕𝟑 𝐇𝐀𝐑𝐃𝐂𝐎𝐑𝐄\n⚡ 𝐕𝐎𝐎𝐍𝐙𝐘 𝐔𝐋𝐓𝐈𝐌𝐀𝐓𝐄 𝐄𝐃𝐈𝐓𝐈𝐎𝐍\n\n⏳ Menunggu otorisasi komandan\n🚫 Akses tidak sah dilarang',
        'welcome_owner': '👑 𝐒𝐔𝐏𝐑𝐄𝐌𝐄 𝐂𝐎𝐌𝐌𝐀𝐍𝐃 𝐂𝐄𝐍𝐓𝐄𝐑\n✨ 𝐕𝐎𝐎𝐍𝐙𝐘 𝐕𝟑 𝐇𝐀𝐑𝐃𝐂𝐎𝐑𝐄 𝐀𝐂𝐂𝐄𝐒𝐒\n\n🎯 Selamat datang kembali Master\n⚡ Ketik /menu untuk deploy',
        'approved': '✅ 𝐀𝐂𝐂𝐄𝐒𝐒 𝐆𝐑𝐀𝐍𝐓𝐄𝐃\n\n⚡ Sistem V3 Hardcore diaktifkan\n🎯 Gunakan /menu untuk dominasi',
        'denied': '❌ 𝐀𝐂𝐂𝐄𝐒𝐒 𝐃𝐄𝐍𝐈𝐄𝐃\n\n🚫 Otorisasi ditolak\n⚡ Hubungi komandan tertinggi',
        'banned': '🚫 𝐏𝐄𝐑𝐌𝐀𝐍𝐄𝐍𝐓𝐋𝐘 𝐁𝐀𝐍𝐍𝐄𝐃\n\n⛔ Anda telah dieliminasi\n💀 Tidak ada banding',
        'menu_title': '✨ 𝐕𝟑 𝐇𝐀𝐑𝐃𝐂𝐎𝐑𝐄 𝐂𝐎𝐍𝐓𝐑𝐎𝐋\n⚡ 𝐕𝐎𝐎𝐍𝐙𝐘 𝐃𝐎𝐌𝐈𝐍𝐀𝐓𝐈𝐎𝐍 𝐓𝐎𝐎𝐋𝐊𝐈𝐓\n\n📍 Pilih senjata Anda',
        'owner_menu_title': '👑 𝐒𝐔𝐏𝐑𝐄𝐌𝐄 𝐃𝐀𝐒𝐇𝐁𝐎𝐀𝐑𝐃\n✨ 𝐌𝐀𝐒𝐓𝐄𝐑 𝐂𝐎𝐍𝐓𝐑𝐎𝐋 𝐕𝟑\n\n⚙️ Semua sistem siap',
        'menu_reg': '🎯 Buat Akun',
        'menu_lang': '🌍 Bahasa',
        'menu_reset': '♻️ Reset',
        'menu_status': '📊 Status',
        'owner_approve': '✅ Persetujuan',
        'owner_users': '👥 Users',
        'owner_domains': '🌐 Domain',
        'gender_title': '✨ 𝐈𝐃𝐄𝐍𝐓𝐈𝐓𝐘 𝐏𝐑𝐎𝐓𝐎𝐂𝐎𝐋\n⚡ Konfigurasi Profil V3\n\n🎭 Pilih gender',
        'gender_male': '👨 Laki-laki',
        'gender_female': '👩 Perempuan',
        'method_title': '✨ 𝐂𝐑𝐄𝐀𝐓𝐈𝐎𝐍 𝐌𝐄𝐓𝐇𝐎𝐃 𝐕𝟑\n⚡ Pemilihan Protokol Hardcore\n\n💡 Pilih senjata Anda',
        'method1': '📱 HP → Email',
        'method2': '✉️ Email Saja',
        'method3': '📞 HP Saja',
        'method4': '📧 HP → TempMail',
        'method5': '🎲 Auto TempMail',
        'method6': '🔥 Gmail → TempMail',
        'method7': '💀 HP Random → Email',
        'method8': '🔥 Hotmail → HP',
        'method9': '💀 Hotmail → TempMail',
        'method10': '🌍 50+ Negara → TempMail',
        'method11': '🌍 50+ Negara → HP',
        'locale_select_title': '🌍 𝐅𝐀𝐂𝐄𝐁𝐎𝐎𝐊 𝐋𝐎𝐂𝐀𝐋𝐄\n⚡ Pemilihan Bahasa Endpoint\n\n📍 Pilih bahasa interface FB',
        'country_select_title': '🌍 𝐂𝐎𝐔𝐍𝐓𝐑𝐘 𝐒𝐄𝐋𝐄𝐂𝐓𝐈𝐎𝐍\n⚡ Region Nomor HP\n\n📍 Pilih negara untuk generate HP',
        'domain_select_title': '📮 𝐓𝐄𝐌𝐏 𝐌𝐀𝐈𝐋 𝐃𝐎𝐌𝐀𝐈𝐍\n⚡ Pemilihan Domain V3\n\n🌐 Pilih domain mail',
        'domain_random': '🎲 Acak',
        'email_prompt': '📧 𝐄𝐌𝐀𝐈𝐋 𝐈𝐍𝐏𝐔𝐓\n\n💬 Masukkan alamat email\n⚠️ /cancel untuk batal',
        'phone_prompt': '📱 𝐏𝐇𝐎𝐍𝐄 𝐈𝐍𝐏𝐔𝐓\n\n💬 Masukkan nomor HP\n🎲 "random" untuk generasi otomatis\n⚠️ /cancel untuk batal',
        'pass_prompt': '🔐 𝐏𝐀𝐒𝐒𝐖𝐎𝐑𝐃 𝐒𝐄𝐓𝐔𝐏\n\n💬 Masukkan password (min 6 karakter)\n🎲 /default untuk generasi otomatis\n⚠️ /cancel untuk batal',
        'processing': '✨ 𝐕𝐎𝐎𝐍𝐙𝐘 𝐕𝟑 𝐇𝐀𝐑𝐃𝐂𝐎𝐑𝐄\n\n🔄 Membuat akun Facebook\n🛡️ Mem-bypass semua keamanan\n⚡ Men-deploy anti-checkpoint\n⏳ Mohon tunggu',
        'error': '❌ 𝐎𝐏𝐄𝐑𝐀𝐓𝐈𝐎𝐍 𝐅𝐀𝐈𝐋𝐄𝐃\n\n💀 Detail Error\n{error}\n\n🔄 /menu untuk coba lagi',
        'success_header': '✨ 𝐀𝐂𝐂𝐎𝐔𝐍𝐓 𝐂𝐑𝐄𝐀𝐓𝐄𝐃\n🎉 𝐃𝐄𝐏𝐋𝐎𝐘𝐌𝐄𝐍𝐓 𝐂𝐎𝐌𝐏𝐋𝐄𝐓𝐄\n',
        'success_name': '\n👤 Nama\n{name}',
        'success_username': '\n\n📧 Username\n{username}',
        'success_password': '\n\n🔑 Password\n{password}',
        'success_uid': '\n\n🆔 UID\n{uid}',
        'success_profile': '\n\n🔗 Profile\n{profile}',
        'success_cookies': '\n\n🍪 Cookies\n{cookies}',
        'success_footer': '\n\n✨ 𝐕𝐎𝐎𝐍𝐙𝐘 𝐕𝟑 | Hardcore Edition',
        'temp_link': '\n\n📬 Temp Mail\n🔗 {link}',
        'reset_done': '✅ 𝐒𝐄𝐒𝐒𝐈𝐎𝐍 𝐑𝐄𝐒𝐄𝐓\n\n♻️ Sesi dibersihkan\n⚡ Siap beraksi',
        'status': '📊 𝐒𝐓𝐀𝐓𝐔𝐒\n\n✅ Disetujui: {approved}\n🚫 Banned: {banned}\n🌍 Bahasa: {lang}\n👤 Gender: {gender}\n\n✨ VOONZY V3 Aktif',
        'back_menu': '◀️ Kembali',
        'cancel': '❌ 𝐂𝐀𝐍𝐂𝐄𝐋𝐋𝐄𝐃\n\n🔙 /menu untuk restart\n✨ VOONZY V3',
        'invalid_pass': '⚠️ 𝐏𝐀𝐒𝐒𝐖𝐎𝐑𝐃 𝐓𝐎𝐎 𝐒𝐇𝐎𝐑𝐓\n\n🔐 Minimal 6 karakter\n🎲 /default untuk auto',
        'no_pending': '✅ 𝐍𝐎 𝐏𝐄𝐍𝐃𝐈𝐍𝐆\n\n⚡ Semua diproses',
        'no_users': '⚠️ 𝐍𝐎 𝐔𝐒𝐄𝐑𝐒\n\n👥 List kosong',
        'domain_list': '🌐 𝐃𝐎𝐌𝐀𝐈𝐍𝐒\n📮 Total: {count}\n\n{domains}\n\n✨ V3',
        'add_domain': '➕ Tambah',
        'delete_domain': '🗑️ Hapus',
        'view_domains': '📋 Lihat Semua',
        'domain_added': '✅ 𝐃𝐎𝐌𝐀𝐈𝐍 𝐀𝐃𝐃𝐄𝐃\n\n⚡ Berhasil ditambahkan\n🌐 V3',
        'domain_deleted': '✅ 𝐃𝐎𝐌𝐀𝐈𝐍 𝐃𝐄𝐋𝐄𝐓𝐄𝐃\n\n⚡ Berhasil dihapus\n🌐 V3',
        'add_domain_prompt': '➕ 𝐀𝐃𝐃 𝐃𝐎𝐌𝐀𝐈𝐍\n\n💬 Masukkan nama domain\n⚠️ /cancel untuk batal',
        'no_domains': '⚠️ 𝐍𝐎 𝐃𝐎𝐌𝐀𝐈𝐍𝐒\n\n🌐 Tambahkan domain dulu',
        'copy_success': '✅ 𝐂𝐎𝐏𝐈𝐄𝐃\n\n📋 {item} tersalin',
        'approval_request': '🔔 𝐍𝐄𝐖 𝐔𝐒𝐄𝐑\n\n👤 Nama: {name}\n🆔 ID: {user_id}\n📱 Username: @{username}\n\n✨ V3',
        'user_approved': '✅ 𝐀𝐏𝐏𝐑𝐎𝐕𝐄𝐃\n\n👤 User ID: {user_id}\n⚡ Akses diberikan',
        'user_denied': '❌ 𝐃𝐄𝐍𝐈𝐄𝐃\n\n👤 User ID: {user_id}\n🚫 Akses ditolak',
        'user_banned': '🚫 𝐁𝐀𝐍𝐍𝐄𝐃\n\n👤 User ID: {user_id}\n⛔ Dibatasi permanen',
        'user_unbanned': '✅ 𝐔𝐍𝐁𝐀𝐍𝐍𝐄𝐃\n\n👤 User ID: {user_id}\n🔓 Pembatasan dihapus',
        'pending_list': '✅ 𝐏𝐄𝐍𝐃𝐈𝐍𝐆\n\n⚡ Pilih untuk setujui/tolak',
        'user_list': '👥 𝐔𝐒𝐄𝐑𝐒\n\n⚡ Klik untuk ban/unban',
        'domain_menu': '🌐 𝐃𝐎𝐌𝐀𝐈𝐍 𝐌𝐀𝐍𝐀𝐆𝐄𝐌𝐄𝐍𝐓\n\n⚡ Kelola domain temp mail',
        'all_domains': '🌐 𝐀𝐋𝐋 𝐃𝐎𝐌𝐀𝐈𝐍𝐒\n📮 Total: {count}\n\n{domains}\n\n✨ V3',
        'delete_domain_menu': '🗑️ 𝐃𝐄𝐋𝐄𝐓𝐄 𝐃𝐎𝐌𝐀𝐈𝐍\n\n⚡ Pilih untuk hapus',
        'lang_changed': '✅ 𝐋𝐀𝐍𝐆𝐔𝐀𝐆𝐄 𝐂𝐇𝐀𝐍𝐆𝐄𝐃\n\n🌍 Bahasa: {lang}\n✨ V3'
    },
    'id': {
        'welcome': '✨ 𝐅𝐁 𝐂𝐑𝐄𝐀𝐓𝐎𝐑 𝐕𝟑 𝐇𝐀𝐑𝐃𝐂𝐎𝐑𝐄\n⚡ 𝐕𝐎𝐎𝐍𝐙𝐘 𝐔𝐋𝐓𝐈𝐌𝐀𝐓𝐄 𝐄𝐃𝐈𝐓𝐈𝐎𝐍\n\n⏳ Menunggu otorisasi komandan\n🚫 Akses tidak sah dilarang',
        'welcome_owner': '👑 𝐒𝐔𝐏𝐑𝐄𝐌𝐄 𝐂𝐎𝐌𝐌𝐀𝐍𝐃 𝐂𝐄𝐍𝐓𝐄𝐑\n✨ 𝐕𝐎𝐎𝐍𝐙𝐘 𝐕𝟑 𝐇𝐀𝐑𝐃𝐂𝐎𝐑𝐄 𝐀𝐂𝐂𝐄𝐒𝐒\n\n🎯 Selamat datang kembali Master\n⚡ Ketik /menu untuk deploy',
        'approved': '✅ 𝐀𝐂𝐂𝐄𝐒𝐒 𝐆𝐑𝐀𝐍𝐓𝐄𝐃\n\n⚡ Sistem V3 Hardcore diaktifkan\n🎯 Gunakan /menu untuk dominasi',
        'denied': '❌ 𝐀𝐂𝐂𝐄𝐒𝐒 𝐃𝐄𝐍𝐈𝐄𝐃\n\n🚫 Otorisasi ditolak\n⚡ Hubungi komandan tertinggi',
        'banned': '🚫 𝐏𝐄𝐑𝐌𝐀𝐍𝐄𝐍𝐓𝐋𝐘 𝐁𝐀𝐍𝐍𝐄𝐃\n\n⛔ Anda telah dieliminasi\n💀 Tidak ada banding',
        'menu_title': '✨ 𝐕𝟑 𝐇𝐀𝐑𝐃𝐂𝐎𝐑𝐄 𝐂𝐎𝐍𝐓𝐑𝐎𝐋\n⚡ 𝐕𝐎𝐎𝐍𝐙𝐘 𝐃𝐎𝐌𝐈𝐍𝐀𝐓𝐈𝐎𝐍 𝐓𝐎𝐎𝐋𝐊𝐈𝐓\n\n📍 Pilih senjata Anda',
        'owner_menu_title': '👑 𝐒𝐔𝐏𝐑𝐄𝐌𝐄 𝐃𝐀𝐒𝐇𝐁𝐎𝐀𝐑𝐃\n✨ 𝐌𝐀𝐒𝐓𝐄𝐑 𝐂𝐎𝐍𝐓𝐑𝐎𝐋 𝐕𝟑\n\n⚙️ Semua sistem siap',
        'menu_reg': '🎯 Buat Akun',
        'menu_lang': '🌍 Bahasa',
        'menu_reset': '♻️ Reset',
        'menu_status': '📊 Status',
        'owner_approve': '✅ Persetujuan',
        'owner_users': '👥 Users',
        'owner_domains': '🌐 Domain',
        'gender_title': '✨ 𝐈𝐃𝐄𝐍𝐓𝐈𝐓𝐘 𝐏𝐑𝐎𝐓𝐎𝐂𝐎𝐋\n⚡ Konfigurasi Profil V3\n\n🎭 Pilih gender',
        'gender_male': '👨 Laki-laki',
        'gender_female': '👩 Perempuan',
        'method_title': '✨ 𝐂𝐑𝐄𝐀𝐓𝐈𝐎𝐍 𝐌𝐄𝐓𝐇𝐎𝐃 𝐕𝟑\n⚡ Pemilihan Protokol Hardcore\n\n💡 Pilih senjata Anda',
        'method1': '📱 HP → Email',
        'method2': '✉️ Email Saja',
        'method3': '📞 HP Saja',
        'method4': '📧 HP → TempMail',
        'method5': '🎲 Auto TempMail',
        'method6': '🔥 Gmail → TempMail',
        'method7': '💀 HP Random → Email',
        'method8': '🔥 Hotmail → HP',
        'method9': '💀 Hotmail → TempMail',
        'method10': '🌍 50+ Negara → TempMail',
        'method11': '🌍 50+ Negara → HP',
        'locale_select_title': '🌍 𝐅𝐀𝐂𝐄𝐁𝐎𝐎𝐊 𝐋𝐎𝐂𝐀𝐋𝐄\n⚡ Pemilihan Bahasa Endpoint\n\n📍 Pilih bahasa interface FB',
        'country_select_title': '🌍 𝐂𝐎𝐔𝐍𝐓𝐑𝐘 𝐒𝐄𝐋𝐄𝐂𝐓𝐈𝐎𝐍\n⚡ Region Nomor HP\n\n📍 Pilih negara untuk generate HP',
        'domain_select_title': '📮 𝐓𝐄𝐌𝐏 𝐌𝐀𝐈𝐋 𝐃𝐎𝐌𝐀𝐈𝐍\n⚡ Pemilihan Domain V3\n\n🌐 Pilih domain mail',
        'domain_random': '🎲 Acak',
        'email_prompt': '📧 𝐄𝐌𝐀𝐈𝐋 𝐈𝐍𝐏𝐔𝐓\n\n💬 Masukkan alamat email\n⚠️ /cancel untuk batal',
        'phone_prompt': '📱 𝐏𝐇𝐎𝐍𝐄 𝐈𝐍𝐏𝐔𝐓\n\n💬 Masukkan nomor HP\n🎲 "random" untuk generasi otomatis\n⚠️ /cancel untuk batal',
        'pass_prompt': '🔐 𝐏𝐀𝐒𝐒𝐖𝐎𝐑𝐃 𝐒𝐄𝐓𝐔𝐏\n\n💬 Masukkan password (min 6 karakter)\n🎲 /default untuk generasi otomatis\n⚠️ /cancel untuk batal',
        'processing': '✨ 𝐕𝐎𝐎𝐍𝐙𝐘 𝐕𝟑 𝐇𝐀𝐑𝐃𝐂𝐎𝐑𝐄\n\n🔄 Membuat akun Facebook\n🛡️ Mem-bypass semua keamanan\n⚡ Men-deploy anti-checkpoint\n⏳ Mohon tunggu',
        'error': '❌ 𝐎𝐏𝐄𝐑𝐀𝐓𝐈𝐎𝐍 𝐅𝐀𝐈𝐋𝐄𝐃\n\n💀 Detail Error\n{error}\n\n🔄 /menu untuk coba lagi',
        'success_header': '✨ 𝐀𝐂𝐂𝐎𝐔𝐍𝐓 𝐂𝐑𝐄𝐀𝐓𝐄𝐃\n🎉 𝐃𝐄𝐏𝐋𝐎𝐘𝐌𝐄𝐍𝐓 𝐂𝐎𝐌𝐏𝐋𝐄𝐓𝐄\n',
        'success_name': '\n👤 Nama\n{name}',
        'success_username': '\n\n📧 Username\n{username}',
        'success_password': '\n\n🔑 Password\n{password}',
        'success_uid': '\n\n🆔 UID\n{uid}',
        'success_profile': '\n\n🔗 Profile\n{profile}',
        'success_cookies': '\n\n🍪 Cookies\n{cookies}',
        'success_footer': '\n\n✨ 𝐕𝐎𝐎𝐍𝐙𝐘 𝐕𝟑 | Hardcore Edition',
        'temp_link': '\n\n📬 Temp Mail\n🔗 {link}',
        'reset_done': '✅ 𝐒𝐄𝐒𝐒𝐈𝐎𝐍 𝐑𝐄𝐒𝐄𝐓\n\n♻️ Sesi dibersihkan\n⚡ Siap beraksi',
        'status': '📊 𝐒𝐓𝐀𝐓𝐔𝐒\n\n✅ Disetujui: {approved}\n🚫 Banned: {banned}\n🌍 Bahasa: {lang}\n👤 Gender: {gender}\n\n✨ VOONZY V3 Aktif',
        'back_menu': '◀️ Kembali',
        'cancel': '❌ 𝐂𝐀𝐍𝐂𝐄𝐋𝐋𝐄𝐃\n\n🔙 /menu untuk restart\n✨ VOONZY V3',
        'invalid_pass': '⚠️ 𝐏𝐀𝐒𝐒𝐖𝐎𝐑𝐃 𝐓𝐎𝐎 𝐒𝐇𝐎𝐑𝐓\n\n🔐 Minimal 6 karakter\n🎲 /default untuk auto',
        'no_pending': '✅ 𝐍𝐎 𝐏𝐄𝐍𝐃𝐈𝐍𝐆\n\n⚡ Semua diproses',
        'no_users': '⚠️ 𝐍𝐎 𝐔𝐒𝐄𝐑𝐒\n\n👥 List kosong',
        'domain_list': '🌐 𝐃𝐎𝐌𝐀𝐈𝐍𝐒\n📮 Total: {count}\n\n{domains}\n\n✨ V3',
        'add_domain': '➕ Tambah',
        'delete_domain': '🗑️ Hapus',
        'view_domains': '📋 Lihat Semua',
        'domain_added': '✅ 𝐃𝐎𝐌𝐀𝐈𝐍 𝐀𝐃𝐃𝐄𝐃\n\n⚡ Berhasil ditambahkan\n🌐 V3',
        'domain_deleted': '✅ 𝐃𝐎𝐌𝐀𝐈𝐍 𝐃𝐄𝐋𝐄𝐓𝐄𝐃\n\n⚡ Berhasil dihapus\n🌐 V3',
        'add_domain_prompt': '➕ 𝐀𝐃𝐃 𝐃𝐎𝐌𝐀𝐈𝐍\n\n💬 Masukkan nama domain\n⚠️ /cancel untuk batal',
        'no_domains': '⚠️ 𝐍𝐎 𝐃𝐎𝐌𝐀𝐈𝐍𝐒\n\n🌐 Tambahkan domain dulu',
        'copy_success': '✅ 𝐂𝐎𝐏𝐈𝐄𝐃\n\n📋 {item} tersalin',
        'approval_request': '🔔 𝐍𝐄𝐖 𝐔𝐒𝐄𝐑\n\n👤 Nama: {name}\n🆔 ID: {user_id}\n📱 Username: @{username}\n\n✨ V3',
        'user_approved': '✅ 𝐀𝐏𝐏𝐑𝐎𝐕𝐄𝐃\n\n👤 User ID: {user_id}\n⚡ Akses diberikan',
        'user_denied': '❌ 𝐃𝐄𝐍𝐈𝐄𝐃\n\n👤 User ID: {user_id}\n🚫 Akses ditolak',
        'user_banned': '🚫 𝐁𝐀𝐍𝐍𝐄𝐃\n\n👤 User ID: {user_id}\n⛔ Dibatasi permanen',
        'user_unbanned': '✅ 𝐔𝐍𝐁𝐀𝐍𝐍𝐄𝐃\n\n👤 User ID: {user_id}\n🔓 Pembatasan dihapus',
        'pending_list': '✅ 𝐏𝐄𝐍𝐃𝐈𝐍𝐆\n\n⚡ Pilih untuk setujui/tolak',
        'user_list': '👥 𝐔𝐒𝐄𝐑𝐒\n\n⚡ Klik untuk ban/unban',
        'domain_menu': '🌐 𝐃𝐎𝐌𝐀𝐈𝐍 𝐌𝐀𝐍𝐀𝐆𝐄𝐌𝐄𝐍𝐓\n\n⚡ Kelola domain temp mail',
        'all_domains': '🌐 𝐀𝐋𝐋 𝐃𝐎𝐌𝐀𝐈𝐍𝐒\n📮 Total: {count}\n\n{domains}\n\n✨ V3',
        'delete_domain_menu': '🗑️ 𝐃𝐄𝐋𝐄𝐓𝐄 𝐃𝐎𝐌𝐀𝐈𝐍\n\n⚡ Pilih untuk hapus',
        'lang_changed': '✅ 𝐋𝐀𝐍𝐆𝐔𝐀𝐆𝐄 𝐂𝐇𝐀𝐍𝐆𝐄𝐃\n\n🌍 Bahasa: {lang}\n✨ V3'
    }
}

# ==================== TELEGRAM HANDLERS ====================

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    user_data = get_user(user.id)
    
    if not user_data:
        add_or_update_user(user.id, user.username or user.first_name or "User")
        user_data = get_user(user.id)
        
        if user.id == OWNER_ID:
            set_approved(user.id, True)
            user_data = get_user(user.id)
        else:
            keyboard = [[
                InlineKeyboardButton("✅ Approve", callback_data=f"approve_{user.id}"),
                InlineKeyboardButton("❌ Deny", callback_data=f"deny_{user.id}")
            ]]
            try:
                texts = LANGUAGES['en']
                await context.bot.send_message(
                    OWNER_ID,
                    texts['approval_request'].format(
                        name=user.first_name,
                        user_id=user.id,
                        username=user.username or 'none'
                    ),
                    reply_markup=InlineKeyboardMarkup(keyboard)
                )
            except Exception as e:
                logger.error(f"Notify owner error: {e}")
            
            await update.message.reply_text(LANGUAGES['en']['welcome'])
            return
    
    if not user_data:
        logger.error(f"User data is None for {user.id}")
        await update.message.reply_text("⚠️ Error loading data. Try /start again.")
        return
    
    texts = LANGUAGES[user_data['language']]
    
    if user_data['banned']:
        await update.message.reply_text(texts['banned'])
        return
    
    if not user_data['approved']:
        if user.id == OWNER_ID:
            set_approved(user.id, True)
            user_data = get_user(user.id)
        else:
            await update.message.reply_text(texts['welcome'])
            return
    
    texts = LANGUAGES[user_data['language']]
    
    if user.id == OWNER_ID:
        keyboard = [
            [InlineKeyboardButton(texts['menu_reg'], callback_data="start_reg")],
            [InlineKeyboardButton(texts['owner_approve'], callback_data="owner_approve"),
             InlineKeyboardButton(texts['owner_users'], callback_data="owner_users")],
            [InlineKeyboardButton(texts['owner_domains'], callback_data="owner_domains")],
            [InlineKeyboardButton(texts['menu_lang'], callback_data="change_lang"),
             InlineKeyboardButton(texts['menu_reset'], callback_data="reset_session")],
            [InlineKeyboardButton(texts['menu_status'], callback_data="my_status")]
        ]
        menu_text = texts['owner_menu_title']
    else:
        keyboard = [
            [InlineKeyboardButton(texts['menu_reg'], callback_data="start_reg")],
            [InlineKeyboardButton(texts['menu_lang'], callback_data="change_lang"),
             InlineKeyboardButton(texts['menu_reset'], callback_data="reset_session")],
            [InlineKeyboardButton(texts['menu_status'], callback_data="my_status")]
        ]
        menu_text = texts['menu_title']
    
    await update.message.reply_text(menu_text, reply_markup=InlineKeyboardMarkup(keyboard))

async def menu_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await start(update, context)

async def button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    user_data = get_user(update.effective_user.id)
    if not user_data:
        return
    
    texts = LANGUAGES[user_data['language']]
    data = query.data
    
    if data.startswith("copy_"):
        parts = data.split("_", 2)
        if len(parts) >= 3:
            action = parts[1]
            uid = parts[2]
            
            result_key = f"result_{uid}"
            if hasattr(context, 'bot_data') and result_key in context.bot_data:
                result = context.bot_data[result_key]
                
                copy_text = ""
                item_name = ""
                
                if action == "uid":
                    copy_text = result.get('uid', '')
                    item_name = "UID"
                elif action == "cookies":
                    copy_text = result.get('cookies', '')
                    item_name = "COOKIES"
                elif action == "all":
                    copy_text = f"""NAME: {result.get('name', '')}
USERNAME: {result.get('username', '')}
PASSWORD: {result.get('password', '')}
UID: {result.get('uid', '')}
PROFILE: {result.get('profile', '')}
COOKIES: {result.get('cookies', '')}"""
                    item_name = "ALL DATA"
                
                await context.bot.send_message(
                    chat_id=update.effective_user.id,
                    text=f"```\n{copy_text}\n```",
                    parse_mode='Markdown'
                )
                
                await query.answer(texts.get('copy_success', '✅ Copied!').format(item=item_name), show_alert=True)
            else:
                await query.answer("⚠️ Data expired. Create new account.", show_alert=True)
        return
    
    if data.startswith("approve_"):
        uid = int(data.split("_")[1])
        set_approved(uid, True)
        target_user = get_user(uid)
        await query.message.reply_text(texts['user_approved'].format(user_id=uid))
        try:
            await context.bot.send_message(uid, LANGUAGES[target_user['language'] if target_user else 'en']['approved'])
        except:
            pass
        return
    
    elif data.startswith("deny_"):
        uid = int(data.split("_")[1])
        await query.message.reply_text(texts['user_denied'].format(user_id=uid))
        try:
            await context.bot.send_message(uid, LANGUAGES['en']['denied'])
        except:
            pass
        return
    
    elif data.startswith("ban_"):
        uid = int(data.split("_")[1])
        set_banned(uid, True)
        keyboard = [[InlineKeyboardButton(texts['back_menu'], callback_data="back_menu")]]
        await query.message.reply_text(texts['user_banned'].format(user_id=uid), reply_markup=InlineKeyboardMarkup(keyboard))
        return
    
    elif data.startswith("unban_"):
        uid = int(data.split("_")[1])
        set_banned(uid, False)
        keyboard = [[InlineKeyboardButton(texts['back_menu'], callback_data="back_menu")]]
        await query.message.reply_text(texts['user_unbanned'].format(user_id=uid), reply_markup=InlineKeyboardMarkup(keyboard))
        return
    
    elif data.startswith("deldomain_"):
        domain = data.split("deldomain_", 1)[1]
        delete_domain(domain)
        keyboard = [[InlineKeyboardButton(texts['back_menu'], callback_data="owner_domains")]]
        await query.message.reply_text(f"{texts['domain_deleted']}\n\n🗑️ {domain}", reply_markup=InlineKeyboardMarkup(keyboard))
        return
    
    elif data.startswith("lang_"):
        lang = data.split("_")[1]
        add_or_update_user(update.effective_user.id, user_data['username'], lang, user_data['gender'])
        keyboard = [[InlineKeyboardButton(LANGUAGES[lang]['back_menu'], callback_data="back_menu")]]
        await query.message.reply_text(LANGUAGES[lang]['lang_changed'].format(lang=lang.upper()), reply_markup=InlineKeyboardMarkup(keyboard))
        return
    
    elif data == "back_menu":
        if update.effective_user.id == OWNER_ID:
            keyboard = [
                [InlineKeyboardButton(texts['menu_reg'], callback_data="start_reg")],
                [InlineKeyboardButton(texts['owner_approve'], callback_data="owner_approve"),
                 InlineKeyboardButton(texts['owner_users'], callback_data="owner_users")],
                [InlineKeyboardButton(texts['owner_domains'], callback_data="owner_domains")],
                [InlineKeyboardButton(texts['menu_lang'], callback_data="change_lang"),
                 InlineKeyboardButton(texts['menu_reset'], callback_data="reset_session")],
                [InlineKeyboardButton(texts['menu_status'], callback_data="my_status")]
            ]
            menu_text = texts['owner_menu_title']
        else:
            keyboard = [
                [InlineKeyboardButton(texts['menu_reg'], callback_data="start_reg")],
                [InlineKeyboardButton(texts['menu_lang'], callback_data="change_lang"),
                 InlineKeyboardButton(texts['menu_reset'], callback_data="reset_session")],
                [InlineKeyboardButton(texts['menu_status'], callback_data="my_status")]
            ]
            menu_text = texts['menu_title']
        
        try:
            await query.edit_message_text(menu_text, reply_markup=InlineKeyboardMarkup(keyboard))
        except:
            await query.message.reply_text(menu_text, reply_markup=InlineKeyboardMarkup(keyboard))
        return
    
    elif data == "change_lang":
        keyboard = [
            [InlineKeyboardButton("🇺🇸 English", callback_data="lang_en")],
            [InlineKeyboardButton("🇮🇩 Indonesia", callback_data="lang_id")],
            [InlineKeyboardButton(texts['back_menu'], callback_data="back_menu")]
        ]
        await query.edit_message_text("🌍 𝐋𝐀𝐍𝐆𝐔𝐀𝐆𝐄\n\n⚡ Pilih bahasa", reply_markup=InlineKeyboardMarkup(keyboard))
        return
    
    elif data == "reset_session":
        clear_session(update.effective_user.id)
        keyboard = [[InlineKeyboardButton(texts['back_menu'], callback_data="back_menu")]]
        await query.edit_message_text(texts['reset_done'], reply_markup=InlineKeyboardMarkup(keyboard))
        return
    
    elif data == "my_status":
        status_text = texts['status'].format(
            approved="Yes" if user_data['approved'] else "Pending",
            banned="Yes" if user_data['banned'] else "No",
            lang=user_data['language'].upper(),
            gender=user_data['gender'].title()
        )
        keyboard = [[InlineKeyboardButton(texts['back_menu'], callback_data="back_menu")]]
        await query.edit_message_text(status_text, reply_markup=InlineKeyboardMarkup(keyboard))
        return
    
    elif data == "owner_approve":
        pendings = get_pending_users()
        if not pendings:
            keyboard = [[InlineKeyboardButton(texts['back_menu'], callback_data="back_menu")]]
            await query.edit_message_text(texts['no_pending'], reply_markup=InlineKeyboardMarkup(keyboard))
            return
        
        keyboard = [[InlineKeyboardButton(f"✅ {p['username']}", callback_data=f"approve_{p['user_id']}"),
                     InlineKeyboardButton(f"❌ {p['username']}", callback_data=f"deny_{p['user_id']}")] for p in pendings[:10]]
        keyboard.append([InlineKeyboardButton(texts['back_menu'], callback_data="back_menu")])
        await query.edit_message_text(texts['pending_list'], reply_markup=InlineKeyboardMarkup(keyboard))
        return
    
    elif data == "owner_users":
        users = get_all_users()
        if not users:
            keyboard = [[InlineKeyboardButton(texts['back_menu'], callback_data="back_menu")]]
            await query.edit_message_text(texts['no_users'], reply_markup=InlineKeyboardMarkup(keyboard))
            return
        
        keyboard = []
        for u in users[:10]:
            status = ("✅" if u['approved'] else "⏳") + (" 🚫" if u['banned'] else "")
            btn_data = f"ban_{u['user_id']}" if not u['banned'] else f"unban_{u['user_id']}"
            keyboard.append([InlineKeyboardButton(f"{status} {u['username']}", callback_data=btn_data)])
        keyboard.append([InlineKeyboardButton(texts['back_menu'], callback_data="back_menu")])
        await query.edit_message_text(texts['user_list'], reply_markup=InlineKeyboardMarkup(keyboard))
        return
    
    elif data == "owner_domains":
        domains = get_domains()
        domain_list = '\n'.join([f"{i+1}. {d}" for i, d in enumerate(domains[:5])])
        if len(domains) > 5:
            domain_list += f"\n... +{len(domains)-5} more"
        
        keyboard = [
            [InlineKeyboardButton(texts['view_domains'], callback_data="view_all_domains")],
            [InlineKeyboardButton(texts['add_domain'], callback_data="add_domain"),
             InlineKeyboardButton(texts['delete_domain'], callback_data="delete_domain_menu")],
            [InlineKeyboardButton(texts['back_menu'], callback_data="back_menu")]
        ]
        await query.edit_message_text(texts['domain_list'].format(count=len(domains), domains=domain_list), reply_markup=InlineKeyboardMarkup(keyboard))
        return
    
    elif data == "view_all_domains":
        domains = get_domains()
        text = texts['all_domains'].format(count=len(domains), domains="\n".join([f"{i+1}. {d}" for i, d in enumerate(domains)]))
        keyboard = [[InlineKeyboardButton(texts['back_menu'], callback_data="owner_domains")]]
        await query.edit_message_text(text[:4000], reply_markup=InlineKeyboardMarkup(keyboard))
        return
    
    elif data == "delete_domain_menu":
        domains = get_domains()
        if not domains:
            keyboard = [[InlineKeyboardButton(texts['back_menu'], callback_data="owner_domains")]]
            await query.edit_message_text(texts['no_domains'], reply_markup=InlineKeyboardMarkup(keyboard))
            return
        
        keyboard = [[InlineKeyboardButton(f"🗑️ {d}", callback_data=f"deldomain_{d}")] for d in domains[:10]]
        keyboard.append([InlineKeyboardButton(texts['back_menu'], callback_data="owner_domains")])
        await query.edit_message_text(texts['delete_domain_menu'], reply_markup=InlineKeyboardMarkup(keyboard))
        return

async def start_registration(update, context):
    query = update.callback_query
    await query.answer()
    user_data = get_user(update.effective_user.id)
    if not user_data:
        return ConversationHandler.END
    texts = LANGUAGES[user_data['language']]
    keyboard = [
        [InlineKeyboardButton(texts['gender_male'], callback_data="gender_male")],
        [InlineKeyboardButton(texts['gender_female'], callback_data="gender_female")],
        [InlineKeyboardButton(texts['back_menu'], callback_data="cancel_reg")]
    ]
    await query.edit_message_text(texts['gender_title'], reply_markup=InlineKeyboardMarkup(keyboard))
    return GENDER_SELECT

async def select_gender(update, context):
    query = update.callback_query
    await query.answer()
    if query.data == "cancel_reg":
        return await cancel_registration(update, context)
    gender = query.data.split("_")[1]
    context.user_data['gender'] = gender
    user_data = get_user(update.effective_user.id)
    add_or_update_user(user_data['user_id'], user_data['username'], user_data['language'], gender)
    texts = LANGUAGES[user_data['language']]
    keyboard = [
        [InlineKeyboardButton(texts['method1'], callback_data="method_1")],
        [InlineKeyboardButton(texts['method2'], callback_data="method_2")],
        [InlineKeyboardButton(texts['method3'], callback_data="method_3")],
        [InlineKeyboardButton(texts['method4'], callback_data="method_4")],
        [InlineKeyboardButton(texts['method5'], callback_data="method_5")],
        [InlineKeyboardButton(texts['method6'], callback_data="method_6")],
        [InlineKeyboardButton(texts['method7'], callback_data="method_7")],
        [InlineKeyboardButton(texts['method8'], callback_data="method_8")],
        [InlineKeyboardButton(texts['method9'], callback_data="method_9")],
        [InlineKeyboardButton(texts['method10'], callback_data="method_10")],
        [InlineKeyboardButton(texts['method11'], callback_data="method_11")],
        [InlineKeyboardButton(texts['back_menu'], callback_data="cancel_reg")]
    ]
    await query.edit_message_text(texts['method_title'], reply_markup=InlineKeyboardMarkup(keyboard))
    return METHOD_SELECT

async def select_method(update, context):
    query = update.callback_query
    await query.answer()
    if query.data == "cancel_reg":
        return await cancel_registration(update, context)
    method = int(query.data.split("_")[1])
    context.user_data['method'] = method
    user_data = get_user(update.effective_user.id)
    texts = LANGUAGES[user_data['language']]
    keyboard = [
        [InlineKeyboardButton(FB_LOCALES['en']['name'], callback_data="locale_en")],
        [InlineKeyboardButton(FB_LOCALES['id']['name'], callback_data="locale_id")],
        [InlineKeyboardButton(FB_LOCALES['es']['name'], callback_data="locale_es")],
        [InlineKeyboardButton(texts['back_menu'], callback_data="cancel_reg")]
    ]
    await query.edit_message_text(texts['locale_select_title'], reply_markup=InlineKeyboardMarkup(keyboard))
    return LOCALE_SELECT

async def select_locale(update, context):
    query = update.callback_query
    await query.answer()
    if query.data == "cancel_reg":
        return await cancel_registration(update, context)
    locale_key = query.data.split("_")[1]
    context.user_data['fb_locale'] = FB_LOCALES[locale_key]['code']
    user_data = get_user(update.effective_user.id)
    texts = LANGUAGES[user_data['language']]
    method = context.user_data.get('method')
    
    if method in [7, 8, 9, 10, 11]:
        keyboard = []
        countries = list(COUNTRY_PHONE_CONFIG.keys())
        for i in range(0, len(countries), 3):
            row = [InlineKeyboardButton(f"🌍 {countries[i+j]}", callback_data=f"country_{countries[i+j]}") 
                   for j in range(3) if i+j < len(countries)]
            keyboard.append(row)
        keyboard.append([InlineKeyboardButton(texts['back_menu'], callback_data="cancel_reg")])
        await query.edit_message_text(texts['country_select_title'], reply_markup=InlineKeyboardMarkup(keyboard))
        return COUNTRY_SELECT
    
    if method in [4, 5, 6, 9, 10]:
        domains = get_domains()
        keyboard = []
        for i in range(0, len(domains), 2):
            row = [InlineKeyboardButton(f"📧 {domains[i][:20]}", callback_data=f"domain_{i}")]
            if i + 1 < len(domains):
                row.append(InlineKeyboardButton(f"📧 {domains[i+1][:20]}", callback_data=f"domain_{i+1}"))
            keyboard.append(row)
        keyboard.append([InlineKeyboardButton(texts['domain_random'], callback_data="domain_random")])
        keyboard.append([InlineKeyboardButton(texts['back_menu'], callback_data="cancel_reg")])
        await query.edit_message_text(texts['domain_select_title'], reply_markup=InlineKeyboardMarkup(keyboard))
        return DOMAIN_SELECT
    
    if method == 2:
        await query.edit_message_text(texts['email_prompt'])
        return EMAIL_INPUT
    if method in [1, 3]:
        await query.edit_message_text(texts['phone_prompt'])
        return PHONE_INPUT
    return COUNTRY_SELECT

async def select_country(update, context):
    query = update.callback_query
    await query.answer()
    if query.data == "cancel_reg":
        return await cancel_registration(update, context)
    country_code = query.data.split("_")[1]
    context.user_data['country_code'] = country_code
    user_data = get_user(update.effective_user.id)
    texts = LANGUAGES[user_data['language']]
    method = context.user_data.get('method')
    
    if method in [9, 10]:
        domains = get_domains()
        keyboard = []
        for i in range(0, len(domains), 2):
            row = [InlineKeyboardButton(f"📧 {domains[i][:20]}", callback_data=f"domain_{i}")]
            if i + 1 < len(domains):
                row.append(InlineKeyboardButton(f"📧 {domains[i+1][:20]}", callback_data=f"domain_{i+1}"))
            keyboard.append(row)
        keyboard.append([InlineKeyboardButton(texts['domain_random'], callback_data="domain_random")])
        keyboard.append([InlineKeyboardButton(texts['back_menu'], callback_data="cancel_reg")])
        await query.edit_message_text(texts['domain_select_title'], reply_markup=InlineKeyboardMarkup(keyboard))
        return DOMAIN_SELECT
    
    if method in [7, 8, 11]:
        await query.edit_message_text(f"🌍 Negara: {country_code}\n\n{texts['pass_prompt']}")
        return PASSWORD_INPUT
    return PASSWORD_INPUT

async def select_domain(update, context):
    query = update.callback_query
    await query.answer()
    if query.data == "cancel_reg":
        return await cancel_registration(update, context)
    user_data = get_user(update.effective_user.id)
    texts = LANGUAGES[user_data['language']]
    
    if query.data == "domain_random":
        context.user_data['selected_domain'] = None
    else:
        domain_index = int(query.data.split("_")[1])
        domains = get_domains()
        context.user_data['selected_domain'] = domains[domain_index]
    
    method = context.user_data.get('method')
    selected = context.user_data.get('selected_domain', 'Random')
    
    if method == 5:
        await query.edit_message_text(f"📧 Domain: {selected}\n\n{texts['pass_prompt']}")
        return PASSWORD_INPUT
    elif method == 6:
        gmail = generate_random_gmail()
        context.user_data['gmail'] = gmail
        await query.edit_message_text(f"🔥 Gmail: {gmail}\n📧 Domain: {selected}\n\n{texts['pass_prompt']}")
        return PASSWORD_INPUT
    elif method in [9, 10]:
        country = context.user_data.get('country_code', 'US')
        await query.edit_message_text(f"🌍 Negara: {country}\n📧 Domain: {selected}\n\n{texts['pass_prompt']}")
        return PASSWORD_INPUT
    else:
        await query.edit_message_text(texts['phone_prompt'])
        return PHONE_INPUT

async def receive_phone(update, context):
    text = update.message.text
    if text == '/cancel':
        return await cancel_registration(update, context)
    user_data = get_user(update.effective_user.id)
    texts = LANGUAGES[user_data['language']]
    
    if text.lower() == 'random':
        country_code = context.user_data.get('country_code', 'ID')
        phone = generate_random_phone_by_country(country_code)
        context.user_data['phone'] = phone
        await update.message.reply_text(f"📱 Generate HP\n\n{phone}\n\n✨ V3")
    else:
        context.user_data['phone'] = text.strip()
    
    method = context.user_data.get('method')
    if method == 1:
        await update.message.reply_text(texts['email_prompt'])
        return EMAIL_INPUT
    else:
        await update.message.reply_text(texts['pass_prompt'])
        return PASSWORD_INPUT

async def receive_email(update, context):
    text = update.message.text
    if text == '/cancel':
        return await cancel_registration(update, context)
    context.user_data['email'] = text.strip()
    user_data = get_user(update.effective_user.id)
    texts = LANGUAGES[user_data['language']]
    await update.message.reply_text(texts['pass_prompt'])
    return PASSWORD_INPUT

async def receive_password(update, context):
    text = update.message.text
    if text == '/cancel':
        return await cancel_registration(update, context)
    user_data = get_user(update.effective_user.id)
    texts = LANGUAGES[user_data['language']]
    
    if text == '/default':
        pass_base = "Password" + str(random.randint(100000, 999999))
    else:
        if len(text.strip()) < 6:
            await update.message.reply_text(texts['invalid_pass'])
            return PASSWORD_INPUT
        pass_base = text.strip()
    
    try:
        if os.path.exists(PROSES_PATH):
            with open(PROSES_PATH, 'rb') as photo:
                sent_msg = await context.bot.send_photo(update.effective_chat.id, photo=photo, caption=texts['processing'])
        else:
            sent_msg = await update.message.reply_text(texts['processing'])
    except:
        sent_msg = await update.message.reply_text(texts['processing'])
    
    user_id = update.effective_user.id
    
    # SIMPAN DATA SEBELUM ASYNC TASK
    method = context.user_data.get('method')
    email = context.user_data.get('email')
    phone = context.user_data.get('phone')
    gender = context.user_data.get('gender')
    selected_domain = context.user_data.get('selected_domain')
    fb_locale = context.user_data.get('fb_locale', 'en_US')
    country_code = context.user_data.get('country_code', 'US')
    language = user_data['language']
    
    async def send_result():
        loop = asyncio.get_event_loop()
        result = await loop.run_in_executor(None, lambda: create_facebook_account_hardcore(
            user_id=user_id,
            method=method,
            email=email,
            phone=phone,
            password=pass_base,
            lang=language,
            gender=gender,
            selected_domain=selected_domain,
            fb_locale=fb_locale,
            country_code=country_code,
            message_id=sent_msg.message_id,
            context=context
        ))
        
        if result.get('success'):
            texts = LANGUAGES[user_data['language']]
            msg = texts['success_header']
            msg += texts['success_name'].format(name=result.get('name', 'N/A'))
            msg += texts['success_username'].format(username=result.get('username', 'N/A'))
            msg += texts['success_password'].format(password=result.get('password', 'N/A'))
            msg += texts['success_uid'].format(uid=result.get('uid', 'N/A'))
            msg += texts['success_profile'].format(profile=result.get('profile', 'N/A'))
            
            cookies_preview = result.get('cookies', 'N/A')[:150] + ("..." if len(result.get('cookies', '')) > 150 else "")
            msg += texts['success_cookies'].format(cookies=cookies_preview)
            
            if result.get('temp_link'):
                msg += texts['temp_link'].format(link=result.get('temp_link', ''))
            
            msg += texts['success_footer']
            msg += f"\n\n🌍 Locale: {result.get('locale', 'N/A')}"
            msg += f"\n🌍 Negara: {result.get('country', 'N/A')}"
            
            keyboard = [
                [InlineKeyboardButton("📋 UID", callback_data=f"copy_uid_{result.get('uid', '')}"),
                 InlineKeyboardButton("📋 Cookies", callback_data=f"copy_cookies_{result.get('uid', '')}")],
                [InlineKeyboardButton("📋 All Data", callback_data=f"copy_all_{result.get('uid', '')}")],
                [InlineKeyboardButton(texts.get('back_menu', '◀️'), callback_data="back_menu")]
            ]
            
            if not hasattr(context, 'bot_data'):
                context.bot_data = {}
            context.bot_data[f"result_{result.get('uid', '')}"] = result
            
            try:
                await context.bot.edit_message_text(chat_id=user_id, message_id=sent_msg.message_id,
                                                   text=msg, reply_markup=InlineKeyboardMarkup(keyboard))
            except:
                await context.bot.send_message(chat_id=user_id, text=msg, reply_markup=InlineKeyboardMarkup(keyboard))
        else:
            msg = texts['error'].format(error=result.get('error', 'Unknown'))
            keyboard = [[InlineKeyboardButton(texts.get('back_menu', '◀️'), callback_data="back_menu")]]
            try:
                await context.bot.edit_message_text(chat_id=user_id, message_id=sent_msg.message_id,
                                                   text=msg, reply_markup=InlineKeyboardMarkup(keyboard))
            except:
                await context.bot.send_message(chat_id=user_id, text=msg, reply_markup=InlineKeyboardMarkup(keyboard))
    
    asyncio.create_task(send_result())
    context.user_data.clear()
    return ConversationHandler.END

async def cancel_registration(update, context):
    user_data = get_user(update.effective_user.id)
    texts = LANGUAGES[user_data['language'] if user_data else 'en']
    context.user_data.clear()
    if update.callback_query:
        await update.callback_query.edit_message_text(texts['cancel'])
    else:
        await update.message.reply_text(texts['cancel'])
    return ConversationHandler.END

async def add_domain_start(update, context):
    query = update.callback_query
    await query.answer()
    user_data = get_user(update.effective_user.id)
    texts = LANGUAGES[user_data['language']]
    await query.edit_message_text(texts['add_domain_prompt'])
    return OWNER_ADD_DOMAIN

async def receive_new_domain(update, context):
    text = update.message.text
    if text == '/cancel':
        user_data = get_user(update.effective_user.id)
        texts = LANGUAGES[user_data['language']]
        await update.message.reply_text(texts['cancel'])
        return ConversationHandler.END
    
    domain = text.strip()
    user_data = get_user(update.effective_user.id)
    texts = LANGUAGES[user_data['language']]
    
    if add_domain(domain):
        await update.message.reply_text(texts['domain_added'])
    else:
        await update.message.reply_text("⚠️ DOMAIN EXISTS\n\nDomain sudah ada di database\n✨ V3")
    
    domains = get_domains()
    domain_list = '\n'.join([f"{i+1}. {d}" for i, d in enumerate(domains[:5])])
    if len(domains) > 5:
        domain_list += f"\n... +{len(domains)-5} more"
    
    keyboard = [
        [InlineKeyboardButton(texts['view_domains'], callback_data="view_all_domains")],
        [InlineKeyboardButton(texts['add_domain'], callback_data="add_domain"),
         InlineKeyboardButton(texts['delete_domain'], callback_data="delete_domain_menu")],
        [InlineKeyboardButton(texts['back_menu'], callback_data="back_menu")]
    ]
    await update.message.reply_text(texts['domain_list'].format(count=len(domains), domains=domain_list), reply_markup=InlineKeyboardMarkup(keyboard))
    return ConversationHandler.END

def main():
    init_db()
    
    application = Application.builder().token(BOT_TOKEN).build()
    
    registration_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(start_registration, pattern="^start_reg$")],
        states={
            GENDER_SELECT: [
                CallbackQueryHandler(select_gender, pattern="^gender_"),
                CallbackQueryHandler(cancel_registration, pattern="^cancel_reg$")
            ],
            METHOD_SELECT: [
                CallbackQueryHandler(select_method, pattern="^method_"),
                CallbackQueryHandler(cancel_registration, pattern="^cancel_reg$")
            ],
            LOCALE_SELECT: [
                CallbackQueryHandler(select_locale, pattern="^locale_"),
                CallbackQueryHandler(cancel_registration, pattern="^cancel_reg$")
            ],
            COUNTRY_SELECT: [
                CallbackQueryHandler(select_country, pattern="^country_"),
                CallbackQueryHandler(cancel_registration, pattern="^cancel_reg$")
            ],
            DOMAIN_SELECT: [
                CallbackQueryHandler(select_domain, pattern="^domain_"),
                CallbackQueryHandler(cancel_registration, pattern="^cancel_reg$")
            ],
            EMAIL_INPUT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, receive_email),
                CommandHandler('cancel', cancel_registration)
            ],
            PHONE_INPUT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, receive_phone),
                CommandHandler('cancel', cancel_registration)
            ],
            PASSWORD_INPUT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, receive_password),
                CommandHandler('cancel', cancel_registration)
            ]
        },
        fallbacks=[
            CommandHandler('cancel', cancel_registration),
            CallbackQueryHandler(cancel_registration, pattern="^cancel_reg$")
        ],
        per_message=False,
        allow_reentry=True
    )
    
    domain_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(add_domain_start, pattern="^add_domain$")],
        states={
            OWNER_ADD_DOMAIN: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, receive_new_domain),
                CommandHandler('cancel', cancel_registration)
            ]
        },
        fallbacks=[CommandHandler('cancel', cancel_registration)],
        per_message=False,
        allow_reentry=True
    )
    
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("menu", menu_command))
    application.add_handler(registration_conv)
    application.add_handler(domain_conv)
    application.add_handler(CallbackQueryHandler(button_handler))

    print("\n" + "="*70)
    print("✨ 𝐕𝐎𝐎𝐍𝐙𝐘 𝐅𝐁 𝐂𝐑𝐄𝐀𝐓𝐎𝐑 𝐕𝟑 - 𝐇𝐀𝐑𝐃𝐂𝐎𝐑𝐄 𝐄𝐃𝐈𝐓𝐈𝐎𝐍 ✨")
    print("⚡ ANTI-CHECKPOINT | MULTI-SESSION | 50+ COUNTRIES")
    print("="*70)
    print(f"✅ Database    : {DB_PATH}")
    print(f"👑 Owner       : {OWNER_ID}")
    print(f"🌐 Domains     : {len(get_domains())}")
    print(f"🌍 Countries   : {len(COUNTRY_PHONE_CONFIG)}")
    print(f"🔥 Concurrent  : 50 sessions")
    print("="*70)
    print("✨ Status      : READY TO DOMINATE")
    print("⚡ Version     : V3 Hardcore Ultra")
    print("🔥 Features    : Anti-Checkpoint | Multi-Locale | Multi-Country")
    print("🛡️ Security    : Advanced Headers | Device Signatures")
    print("="*70 + "\n")
    
    logger.info("✨ VoonzyV3 Hardcore - Bot started")
    application.run_polling(
        allowed_updates=Update.ALL_TYPES,
        drop_pending_updates=True
    )

if __name__ == "__main__":
    main()
